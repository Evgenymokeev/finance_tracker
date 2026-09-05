from io import BytesIO

from django.contrib.auth.models import User
from django.core.files.uploadedfile import SimpleUploadedFile

from openpyxl import load_workbook

from rest_framework import status
from rest_framework.test import APITestCase

from categories.models import Category
from expenses.models import Expense


class ExpenseAPITest(APITestCase):

    def setUp(self):
        self.user = User.objects.create_user(
            username="testuser",
            password="testpass123",
        )

        self.category = Category.objects.create(
            user=self.user,
            name="Еда",
        )

        self.client.force_authenticate(
            user=self.user
        )

    # =========================================================
    # CREATE
    # =========================================================

    def test_create_expense(self):
        data = {
            "title": "Кофе",
            "amount": "80.00",
            "category": self.category.id,
            "date": "2026-07-23",
            "description": "Latte",
        }

        response = self.client.post(
            "/api/v1/expenses/",
            data,
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_201_CREATED,
        )

        self.assertEqual(
            Expense.objects.count(),
            1,
        )

        expense = Expense.objects.first()

        self.assertEqual(
            expense.user,
            self.user,
        )

    def test_unauthorized_user_cannot_create_expense(self):
        self.client.force_authenticate(user=None)

        data = {
            "title": "Coffee",
            "amount": "120.00",
            "category": self.category.id,
            "date": "2026-07-24",
            "description": "Latte",
        }

        response = self.client.post(
            "/api/v1/expenses/",
            data,
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_401_UNAUTHORIZED,
        )

    def test_cannot_create_expense_with_foreign_category(self):
        second_user = User.objects.create_user(
            username="seconduser",
            password="testpass123",
        )

        foreign_category = Category.objects.create(
            user=second_user,
            name="Чужая категория",
        )

        data = {
            "title": "Кофе",
            "amount": "80.00",
            "category": foreign_category.id,
            "date": "2026-07-23",
            "description": "Latte",
        }

        response = self.client.post(
            "/api/v1/expenses/",
            data,
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_400_BAD_REQUEST,
        )

        self.assertEqual(
            Expense.objects.count(),
            0,
        )

    # =========================================================
    # LIST / USER ISOLATION
    # =========================================================

    def test_user_sees_only_own_expenses(self):
        second_user = User.objects.create_user(
            username="seconduser",
            password="testpass123",
        )

        second_category = Category.objects.create(
            user=second_user,
            name="Транспорт",
        )

        Expense.objects.create(
            user=self.user,
            title="Кофе",
            amount="80.00",
            category=self.category,
            date="2026-07-23",
        )

        Expense.objects.create(
            user=second_user,
            title="Автобус",
            amount="40.00",
            category=second_category,
            date="2026-07-23",
        )

        response = self.client.get(
            "/api/v1/expenses/",
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_200_OK,
        )

        self.assertEqual(
            len(response.data["results"]),
            1,
        )

        self.assertEqual(
            response.data["results"][0]["title"],
            "Кофе",
        )

    # =========================================================
    # FILTERS
    # =========================================================

    def test_filter_expenses_by_category(self):
        transport = Category.objects.create(
            user=self.user,
            name="Транспорт",
        )

        Expense.objects.create(
            user=self.user,
            title="Кофе",
            amount="80.00",
            category=self.category,
            date="2026-08-01",
        )

        Expense.objects.create(
            user=self.user,
            title="Такси",
            amount="250.00",
            category=transport,
            date="2026-08-02",
        )

        response = self.client.get(
            f"/api/v1/expenses/?category={self.category.id}"
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_200_OK,
        )

        self.assertEqual(
            len(response.data["results"]),
            1,
        )

        self.assertEqual(
            response.data["results"][0]["title"],
            "Кофе",
        )

    def test_filter_expenses_by_date_range(self):
        Expense.objects.create(
            user=self.user,
            title="Старый расход",
            amount="50.00",
            category=self.category,
            date="2026-07-01",
        )

        Expense.objects.create(
            user=self.user,
            title="Расход в диапазоне",
            amount="100.00",
            category=self.category,
            date="2026-08-10",
        )

        Expense.objects.create(
            user=self.user,
            title="Будущий расход",
            amount="150.00",
            category=self.category,
            date="2026-09-01",
        )

        response = self.client.get(
            "/api/v1/expenses/"
            "?date_from=2026-08-01"
            "&date_to=2026-08-31"
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_200_OK,
        )

        self.assertEqual(
            len(response.data["results"]),
            1,
        )

        self.assertEqual(
            response.data["results"][0]["title"],
            "Расход в диапазоне",
        )

    def test_filter_expenses_by_amount_range(self):
        Expense.objects.create(
            user=self.user,
            title="Дешёвый расход",
            amount="50.00",
            category=self.category,
            date="2026-08-01",
        )

        Expense.objects.create(
            user=self.user,
            title="Расход в диапазоне",
            amount="100.00",
            category=self.category,
            date="2026-08-02",
        )

        Expense.objects.create(
            user=self.user,
            title="Дорогой расход",
            amount="500.00",
            category=self.category,
            date="2026-08-03",
        )

        response = self.client.get(
            "/api/v1/expenses/"
            "?amount_from=80"
            "&amount_to=150"
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_200_OK,
        )

        self.assertEqual(
            len(response.data["results"]),
            1,
        )

        self.assertEqual(
            response.data["results"][0]["title"],
            "Расход в диапазоне",
        )

    def test_filter_expenses_combined(self):
        transport = Category.objects.create(
            user=self.user,
            name="Транспорт",
        )

        Expense.objects.create(
            user=self.user,
            title="Кофе",
            amount="80.00",
            category=self.category,
            date="2026-08-05",
        )

        Expense.objects.create(
            user=self.user,
            title="Такси",
            amount="200.00",
            category=transport,
            date="2026-08-10",
        )

        Expense.objects.create(
            user=self.user,
            title="Обед",
            amount="120.00",
            category=self.category,
            date="2026-09-01",
        )

        response = self.client.get(
            "/api/v1/expenses/"
            f"?category={self.category.id}"
            "&date_from=2026-08-01"
            "&date_to=2026-08-31"
            "&amount_from=50"
            "&amount_to=100"
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_200_OK,
        )

        self.assertEqual(
            len(response.data["results"]),
            1,
        )

        self.assertEqual(
            response.data["results"][0]["title"],
            "Кофе",
        )

    # =========================================================
    # CSV EXPORT
    # =========================================================

    def test_export_expenses_to_csv(self):
        Expense.objects.create(
            user=self.user,
            title="Кофе",
            amount="80.00",
            category=self.category,
            date="2026-07-23",
            description="Latte",
        )

        response = self.client.get(
            "/api/v1/expenses/export/",
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_200_OK,
        )

        self.assertEqual(
            response["Content-Type"],
            "text/csv",
        )

        self.assertIn(
            "attachment",
            response["Content-Disposition"],
        )

        content = response.content.decode("utf-8")

        self.assertIn(
            "Date,Title,Category,Amount,Description",
            content,
        )

        self.assertIn(
            "Кофе",
            content,
        )

        self.assertIn(
            "Еда",
            content,
        )

    def test_export_contains_only_own_expenses(self):
        second_user = User.objects.create_user(
            username="seconduser",
            password="testpass123",
        )

        second_category = Category.objects.create(
            user=second_user,
            name="Транспорт",
        )

        Expense.objects.create(
            user=self.user,
            title="Мой кофе",
            amount="80.00",
            category=self.category,
            date="2026-08-01",
            description="Latte",
        )

        Expense.objects.create(
            user=second_user,
            title="Чужой автобус",
            amount="40.00",
            category=second_category,
            date="2026-08-01",
            description="Bus",
        )

        response = self.client.get(
            "/api/v1/expenses/export/",
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_200_OK,
        )

        content = response.content.decode("utf-8")

        self.assertIn(
            "Мой кофе",
            content,
        )

        self.assertNotIn(
            "Чужой автобус",
            content,
        )

    # =========================================================
    # EXCEL EXPORT
    # =========================================================

    def test_export_expenses_to_excel(self):
        Expense.objects.create(
            user=self.user,
            title="Кофе",
            amount="80.00",
            category=self.category,
            date="2026-08-01",
            description="Latte",
        )

        response = self.client.get(
            "/api/v1/expenses/export-excel/",
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_200_OK,
        )

        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        self.assertIn(
            "attachment",
            response["Content-Disposition"],
        )

        self.assertIn(
            "expenses.xlsx",
            response["Content-Disposition"],
        )

        self.assertGreater(
            len(response.content),
            0,
        )

    def test_export_excel_contains_only_own_expenses(self):
        second_user = User.objects.create_user(
            username="seconduser",
            password="testpass123",
        )

        second_category = Category.objects.create(
            user=second_user,
            name="Транспорт",
        )

        Expense.objects.create(
            user=self.user,
            title="Мой кофе",
            amount="80.00",
            category=self.category,
            date="2026-08-01",
            description="Latte",
        )

        Expense.objects.create(
            user=second_user,
            title="Чужой автобус",
            amount="40.00",
            category=second_category,
            date="2026-08-01",
            description="Bus",
        )

        response = self.client.get(
            "/api/v1/expenses/export-excel/",
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_200_OK,
        )

        workbook = load_workbook(
            BytesIO(response.content)
        )

        worksheet = workbook.active

        rows = list(
            worksheet.iter_rows(
                values_only=True
            )
        )

        self.assertEqual(
            rows[0],
            (
                "Date",
                "Title",
                "Category",
                "Amount",
                "Description",
            ),
        )

        self.assertEqual(
            len(rows),
            2,
        )

        self.assertEqual(
            rows[1][1],
            "Мой кофе",
        )

        self.assertNotIn(
            "Чужой автобус",
            [row[1] for row in rows],
        )

    def test_export_excel_respects_filters(self):
        Expense.objects.create(
            user=self.user,
            title="Август",
            amount="100.00",
            category=self.category,
            date="2026-08-10",
            description="August expense",
        )

        Expense.objects.create(
            user=self.user,
            title="Июль",
            amount="200.00",
            category=self.category,
            date="2026-07-10",
            description="July expense",
        )

        response = self.client.get(
            "/api/v1/expenses/export-excel/",
            {
                "date_from": "2026-08-01",
                "date_to": "2026-08-31",
            },
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_200_OK,
        )

        workbook = load_workbook(
            BytesIO(response.content)
        )

        worksheet = workbook.active

        rows = list(
            worksheet.iter_rows(
                values_only=True
            )
        )

        self.assertEqual(
            len(rows),
            2,
        )

        self.assertEqual(
            rows[1][1],
            "Август",
        )

    def test_unauthorized_user_cannot_export_excel(self):
        self.client.force_authenticate(user=None)

        response = self.client.get(
            "/api/v1/expenses/export-excel/",
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_401_UNAUTHORIZED,
        )

    # =========================================================
    # CSV IMPORT
    # =========================================================

    def test_import_expenses_from_csv(self):
        csv_content = (
            "date,title,category,amount,description\n"
            "2026-08-01,Coffee,Еда,80,Latte\n"
            "2026-08-02,Taxi,Еда,250,\n"
        )

        csv_file = SimpleUploadedFile(
            "expenses.csv",
            csv_content.encode("utf-8"),
            content_type="text/csv",
        )

        response = self.client.post(
            "/api/v1/expenses/import/",
            {"file": csv_file},
            format="multipart",
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_200_OK,
        )

        self.assertEqual(
            response.data["created"],
            2,
        )

        self.assertEqual(
            response.data["errors"],
            [],
        )

        self.assertEqual(
            Expense.objects.count(),
            2,
        )

    def test_import_expenses_with_unknown_category(self):
        csv_content = (
            "date,title,category,amount,description\n"
            "2026-08-01,Coffee,Несуществующая,80,Latte\n"
        )

        csv_file = SimpleUploadedFile(
            "expenses.csv",
            csv_content.encode("utf-8"),
            content_type="text/csv",
        )

        response = self.client.post(
            "/api/v1/expenses/import/",
            {"file": csv_file},
            format="multipart",
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_200_OK,
        )

        self.assertEqual(
            response.data["created"],
            0,
        )

        self.assertEqual(
            len(response.data["errors"]),
            1,
        )

        self.assertEqual(
            response.data["errors"][0]["row"],
            2,
        )

        self.assertIn(
            "does not exist",
            response.data["errors"][0]["error"],
        )

        self.assertEqual(
            Expense.objects.count(),
            0,
        )

    def test_import_expenses_partial_success(self):
        csv_content = (
            "date,title,category,amount,description\n"
            "2026-08-01,Coffee,Еда,80,Latte\n"
            "2026-08-02,Taxi,Несуществующая,250,\n"
            "2026-08-03,Lunch,Еда,150,\n"
        )

        csv_file = SimpleUploadedFile(
            "expenses.csv",
            csv_content.encode("utf-8"),
            content_type="text/csv",
        )

        response = self.client.post(
            "/api/v1/expenses/import/",
            {"file": csv_file},
            format="multipart",
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_200_OK,
        )

        self.assertEqual(
            response.data["created"],
            2,
        )

        self.assertEqual(
            len(response.data["errors"]),
            1,
        )

        self.assertEqual(
            response.data["errors"][0]["row"],
            3,
        )

        self.assertEqual(
            Expense.objects.count(),
            2,
        )

    def test_import_expenses_missing_columns(self):
        csv_content = (
            "date,title,amount\n"
            "2026-08-01,Coffee,80\n"
        )

        csv_file = SimpleUploadedFile(
            "expenses.csv",
            csv_content.encode("utf-8"),
            content_type="text/csv",
        )

        response = self.client.post(
            "/api/v1/expenses/import/",
            {"file": csv_file},
            format="multipart",
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_200_OK,
        )

        self.assertEqual(
            response.data["created"],
            0,
        )

        self.assertEqual(
            len(response.data["errors"]),
            1,
        )

        self.assertIn(
            "Missing required columns",
            response.data["errors"][0]["error"],
        )

        self.assertEqual(
            Expense.objects.count(),
            0,
        )

    def test_unauthorized_user_cannot_import_expenses(self):
        self.client.force_authenticate(user=None)

        csv_content = (
            "date,title,category,amount,description\n"
            "2026-08-01,Coffee,Еда,80,Latte\n"
        )

        csv_file = SimpleUploadedFile(
            "expenses.csv",
            csv_content.encode("utf-8"),
            content_type="text/csv",
        )

        response = self.client.post(
            "/api/v1/expenses/import/",
            {"file": csv_file},
            format="multipart",
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_401_UNAUTHORIZED,
        )

        self.assertEqual(
            Expense.objects.count(),
            0,
        )

    # =========================================================
    # CSV IMPORT — FILE VALIDATION
    # =========================================================

    def test_import_invalid_utf8_csv(self):
        csv_file = SimpleUploadedFile(
            "expenses.csv",
            b"\xff\xfe\xfd",
            content_type="text/csv",
        )

        response = self.client.post(
            "/api/v1/expenses/import/",
            {"file": csv_file},
            format="multipart",
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_200_OK,
        )

        self.assertEqual(
            response.data["created"],
            0,
        )

        self.assertEqual(
            len(response.data["errors"]),
            1,
        )

        self.assertEqual(
            response.data["errors"][0]["row"],
            1,
        )

        self.assertEqual(
            response.data["errors"][0]["error"],
            "CSV file must be encoded in UTF-8.",
        )

        self.assertEqual(
            Expense.objects.count(),
            0,
        )

    def test_import_empty_csv(self):
        csv_file = SimpleUploadedFile(
            "expenses.csv",
            b"",
            content_type="text/csv",
        )

        response = self.client.post(
            "/api/v1/expenses/import/",
            {"file": csv_file},
            format="multipart",
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_400_BAD_REQUEST,
        )

        self.assertEqual(
            Expense.objects.count(),
            0,
        )

    # =========================================================
    # CSV IMPORT — REQUIRED VALUES
    # =========================================================

    def test_import_expense_without_title(self):
        csv_content = (
            "date,title,category,amount,description\n"
            "2026-08-01,,Еда,80,Latte\n"
        )

        csv_file = SimpleUploadedFile(
            "expenses.csv",
            csv_content.encode("utf-8"),
            content_type="text/csv",
        )

        response = self.client.post(
            "/api/v1/expenses/import/",
            {"file": csv_file},
            format="multipart",
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_200_OK,
        )

        self.assertEqual(
            response.data["created"],
            0,
        )

        self.assertEqual(
            response.data["errors"][0]["row"],
            2,
        )

        self.assertEqual(
            response.data["errors"][0]["error"],
            "Title is required.",
        )

    def test_import_expense_without_category(self):
        csv_content = (
            "date,title,category,amount,description\n"
            "2026-08-01,Coffee,,80,Latte\n"
        )

        csv_file = SimpleUploadedFile(
            "expenses.csv",
            csv_content.encode("utf-8"),
            content_type="text/csv",
        )

        response = self.client.post(
            "/api/v1/expenses/import/",
            {"file": csv_file},
            format="multipart",
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_200_OK,
        )

        self.assertEqual(
            response.data["created"],
            0,
        )

        self.assertEqual(
            response.data["errors"][0]["row"],
            2,
        )

        self.assertEqual(
            response.data["errors"][0]["error"],
            "Category is required.",
        )

    def test_import_expense_without_amount(self):
        csv_content = (
            "date,title,category,amount,description\n"
            "2026-08-01,Coffee,Еда,,Latte\n"
        )

        csv_file = SimpleUploadedFile(
            "expenses.csv",
            csv_content.encode("utf-8"),
            content_type="text/csv",
        )

        response = self.client.post(
            "/api/v1/expenses/import/",
            {"file": csv_file},
            format="multipart",
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_200_OK,
        )

        self.assertEqual(
            response.data["created"],
            0,
        )

        self.assertEqual(
            response.data["errors"][0]["row"],
            2,
        )

        self.assertEqual(
            response.data["errors"][0]["error"],
            "Amount is required.",
        )

    def test_import_expense_without_date(self):
        csv_content = (
            "date,title,category,amount,description\n"
            ",Coffee,Еда,80,Latte\n"
        )

        csv_file = SimpleUploadedFile(
            "expenses.csv",
            csv_content.encode("utf-8"),
            content_type="text/csv",
        )

        response = self.client.post(
            "/api/v1/expenses/import/",
            {"file": csv_file},
            format="multipart",
        )

        self.assertEqual(
            response.status_code,
            status.HTTP_200_OK,
        )

        self.assertEqual(
            response.data["created"],
            0,
        )

        self.assertEqual(
            response.data["errors"][0]["row"],
            2,
        )

        self.assertEqual(
            response.data["errors"][0]["error"],
            "Date is required.",
        )
